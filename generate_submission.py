"""
Amex Campus Challenge 2026 — Final Submission Generator
========================================================
Produces final_submission.xlsx ready for Unstop upload.

Two sheets are filled:
  1. Predictions         — ID + Prediction (profit score) for all 500,000 rows
  2. Profitability Framework — 10 text explanation fields for Amex evaluators

Run:
    python generate_submission.py

Output:
    eda_outputs/final_submission.xlsx
"""

import warnings
import numpy as np
import pandas as pd
import openpyxl

warnings.filterwarnings("ignore")

# ── paths ──────────────────────────────────────────────────────────────────────
RANKED_PATH   = "eda_outputs/data_ranked.csv"
IMPUTED_PATH  = "eda_outputs/data_imputed.csv"
TEMPLATE_PATH = "6a3cb64c7cae4_campus_challenge_r1_submission_template.xlsx"
OUT_PATH      = "eda_outputs/final_submission.xlsx"

SEP = "=" * 70

def section(title):
    print(f"\n{SEP}\n  {title}\n{SEP}")


# ══════════════════════════════════════════════════════════════════════════════
# v5 EFFECTIVE WEIGHTS  (same as validation_harness.py — do not change here)
# ══════════════════════════════════════════════════════════════════════════════
EFFECTIVE_WEIGHTS = {
    "f1":  +0.20,
    "f5":  +0.22,
    "f6":  +0.05,
    "f10": +0.03,
    "f7":  +0.03,
    "f8":  +0.015,
    "f9":  +0.015,
    "f17": +0.01,
    "f18": +0.01,
    "f11": -0.26,
    "f21": -0.10,
    "f15": -0.09,
    "f4":  -0.07,
    "f13": -0.07,
    "f14": -0.05,
    "f2":  -0.02,
}

# ══════════════════════════════════════════════════════════════════════════════
# PROFITABILITY FRAMEWORK TEXT  (shown to Amex evaluators — keep concise)
# ══════════════════════════════════════════════════════════════════════════════
FRAMEWORK = {
    "Variables Used": (
        "Revenue: f1 (Avg Revolve Balance), f5 (Total Spend), f6 (Airlines Spend), f7 (Other Spend), "
        "f8 (Entertainment Spend), f9 (Lodging Spend), f10 (Dining Spend), f17 (Total Lend Line), "
        "f18 (Consumer Lend Line). "
        "Cost: f11 (Avg Risk Score), f21 (Rewards Points Redeemed), f15 (Cab Benefits Used months), "
        "f4 (Rewards Points Balance), f13 (Lounge Access Count), f14 (Airline Credits Used), "
        "f2 (Cancellation Calls). "
        "Hard override: f3 (Collections Calls) — any customer with f3>0 is bottom-ranked as a near-certain loss. "
        "Excluded (weight 0): f16 (hard-capped at $64.40, near-binary in top quartile — structural inversion), "
        "f12/f19/f20/f22 (engagement or profile context, no direct P&L impact), f23 (87.8% null — dropped)."
    ),

    "Profitability Equation": (
        "All coefficients below are EFFECTIVE weights (desired sorting power over the full population). "
        "Raw weights are back-solved as: raw = effective / coverage, where coverage = 1 - zero_fraction. "
        "\n\n"
        "Score = "
        "+ 0.22 x rank(f5)   [Total Spend — interchange anchor] "
        "+ 0.20 x rank(f1)   [Revolve Balance — NIM engine] "
        "+ 0.05 x rank(f6)   [Airlines Spend — premium interchange] "
        "+ 0.03 x rank(f10)  [Dining Spend — premium interchange] "
        "+ 0.03 x rank(f7)   [Other Spend — generic interchange] "
        "+ 0.015 x rank(f8)  [Entertainment Spend] "
        "+ 0.015 x rank(f9)  [Lodging Spend] "
        "+ 0.01 x rank(f17)  [Total Lend Line] "
        "+ 0.01 x rank(f18)  [Consumer Lend Line] "
        "- 0.26 x rank(f11)  [Risk Score — Expected Credit Loss] "
        "- 0.10 x rank(f21)  [Points Redeemed — realized cost] "
        "- 0.09 x rank(f15)  [Cab Months — $15/month partner cost] "
        "- 0.07 x rank(f4)   [Rewards Balance — IFRS 15 liability] "
        "- 0.07 x rank(f13)  [Lounge Visits — $24-32/visit] "
        "- 0.05 x rank(f14)  [Airline Credit Used] "
        "- 0.02 x rank(f2)   [Cancellation Calls — servicing overhead] "
        "\n\n"
        "f3>0 customers sunk to the bottom of the ordinary score range (no -999 sentinel). "
        "Effective revenue total: +0.580 | Cost: -0.660 | Net: -0.080"
    ),

    "Prediction Logic": (
        "Step 1 — Imputation: structured null groups are zero-filled (MNAR interpretation — missing = "
        "no economic activity, not unknown). f5 and f11 are median-imputed (low null rate, genuinely unknown). "
        "f23 dropped (87.8% null). "
        "Step 2 — Zero-floor rank-transform: customers with value=0 receive rank 0.0 exactly (hard floor). "
        "Non-zero values ranked proportionally within (0,1] among the active population only. "
        "f7 uses standard rank (has real negative values from refunds that must rank below zero-spend). "
        "Step 3 — Coverage back-solve: raw weight = effective weight / (1 - zero_fraction) per feature. "
        "Step 4 — Weighted sum of rank-transformed features using raw weights. "
        "Step 5 — f3>0 customers (54,304 = 11% of population) are placed at the bottom of the ordinary "
        "score range, below all eligible customers. Top 20% is drawn entirely from the remaining 89%. "
        "Step 6 — Scores exported sorted by ID for submission."
    ),

    "Variable Selection Logic": (
        "Each of the 23 features was classified as Revenue, Cost, or Context using Premier Card P&L economics. "
        "Revenue features included: revolving interest (f1), total and category interchange spend (f5-f10), "
        "lend capacity as a Plan-It trust signal (f17, f18). "
        "Cost features included: expected credit loss (f11), realized rewards redemption (f21), "
        "benefit utilization costs (f13 lounge ~$30/visit, f14 airline credit, f15 cab $15/month), "
        "deferred rewards liability (f4, IFRS 15), servicing overhead (f2). "
        "f3 treated as hard override not a weight — a customer with active collection calls cannot be profitable "
        "regardless of spend. "
        "f16 excluded: hard-capped at $64.40 (75th percentile = maximum), near-binary above the 75th pct — "
        "it cannot differentiate customers in the top quartile, and the top 20% (high spenders) also use this "
        "credit, making inversion impossible to resolve by any weight. "
        "f12 (logins), f19 (supplementary accounts), f20 (charge cards), f22 (emails opened): "
        "engagement/profile proxies — predict churn or tenure, not P&L. "
        "f23: 87.8% null, near-zero discriminatory power — dropped entirely."
    ),

    "Coefficient/Weight Derivation": (
        "All weights grounded in publicly cited card economics — not fitted to a label (there is none). "
        "f5 (+0.22 eff): Amex premium merchant processing ~2.50%+$0.10/txn; highest coverage (93.5%) "
        "makes this the strongest single lever. "
        "f1 (+0.20 eff): revolving APR 19.99-29.99% — a dollar of revolve yields ~10x a dollar of spend revenue. "
        "f6 (+0.05 eff): airline MCCs (3000-3299, 4511) at 2.30-3.30% — richest interchange tier. "
        "f10 (+0.03 eff): dining MCCs (5812, 5814) at 1.85-2.75%+$0.10. "
        "f7 (+0.03 eff): retail/other 1.43-2.40%+$0.10. "
        "f11 (-0.26 eff): heaviest cost — one default wipes the interchange from dozens of transactors "
        "(Basel III Expected Loss = PD x EAD). "
        "f21 (-0.10 eff): realized redemption at ~1.0c/pt (Amex Travel portal flights); kept below f11 "
        "because credit risk > rewards cost in P&L priority. "
        "f15 (-0.09 eff): Uber Cash structured at $15/month. "
        "f4 (-0.07 eff): IFRS 15 deferred liability, discounted 15-20% breakage. "
        "f13 (-0.07 eff): Priority Pass partner fee ~USD 32/guest visit. "
        "f14 (-0.05 eff): airline fee credit, hard-capped at $200/year. "
        "f2 (-0.02 eff): minor — retention call servicing overhead."
    ),

    "Feature Transformations": (
        "All features rank-transformed to [0,1] percentile scale before weighting. "
        "Rationale: (1) evaluation metric is rank-based (top-20% overlap) so rank-transform aligns "
        "inputs directly with the scoring mechanic; (2) robust to extreme right-skew (f7 max ~147,000 "
        "vs f11 max 0.33 — raw summation would let f7 silently dominate); (3) no distributional assumptions. "
        "\n"
        "Zero-floor rank (critical): standard rank assigns tied zeros the average rank of the tie block. "
        "For f21 (72% zeros), standard rank gives zeros rank ~0.36 — manufacturing phantom signal for "
        "customers with no redemption activity. Zero-floor fix: all post-imputation zero values -> rank 0.0 "
        "(hard floor); non-zero values ranked proportionally within (0,1] among the active-only population. "
        "Applied to all features where >20% of values are zero. "
        "Exception: f7 uses standard rank because its negative values (refunds, min -275) are economically "
        "informative and must rank below zero-spend customers."
    ),

    "Business Logic": (
        "Framework: Net Profitability = Gross Revenue - Direct Variable Costs - Expected Credit Loss. "
        "\n"
        "Revenue streams modeled: "
        "(a) Interchange / discount revenue — merchant fee on every dollar spent (f5, f6-f10, "
        "category-differentiated by MCC interchange rate). "
        "(b) Interest / NIM revenue — revolving balance at 20-30% APR (f1) and lend line capacity (f17, f18). "
        "\n"
        "Cost streams modeled: "
        "(a) Expected Credit Loss — risk score x exposure (f11); collections calls as hard knockout (f3). "
        "(b) Rewards cost — realized redemptions at ~1c/pt (f21) and deferred balance liability (f4). "
        "(c) Benefit cost — lounge visits at ~$30/visit (f13), airline credit (f14), cab credit $15/month (f15). "
        "(d) Servicing — cancellation/retention calls (f2). "
        "\n"
        "Core economic insight: the most profitable cardmember is high-spending, moderately-revolving, "
        "low-risk, and does NOT fully exploit the benefits. A high-risk customer who maxes every credit "
        "and redeems all points is a loss-maker even if they look 'engaged.' Benefit utilization and risk "
        "terms must subtract — any equation that rewards high benefit usage has inverted the economics."
    ),

    "Assumptions": (
        "1. Interchange rates: Amex premium processing 2.50%+$0.10/txn average; "
        "airlines 2.30-3.30%, dining 1.85-2.75%, retail/other 1.43-2.40% "
        "(Amex OptBlue / MCC wholesale discount rate schedules). "
        "2. Revolving APR: 19.99-29.99% (standard premium card product disclosures). "
        "3. Points redemption value: ~1.0c/pt via Amex Travel portal flight redemptions "
        "(Membership Rewards program terms; Pay-with-Points ~0.7c/pt). "
        "4. Lounge issuer cost: ~$24-32/visit (Priority Pass USD 32 guest fee per additional visit). "
        "5. Cab credit cost: $15/month structured benefit + $20 December bonus "
        "(Amex Platinum Uber Cash benefit terms). "
        "6. IFRS 15 breakage rate on unredeemed points: 15-20% "
        "(standard deferred revenue recognition). "
        "7. Expected Credit Loss modeled as Basel III PD x EAD. "
        "8. f4/f21 null segment (51.4%): zero-filled — interpreted as dormant/masked customers "
        "with no recorded rewards activity in the data window. NOT 'not enrolled' — all Premier "
        "cardmembers earn 1x Membership Rewards points by default on every purchase. "
        "9. f5 and f6-f10 are independently masked/scaled (Spearman r=0.09 between f5 and sum(f6:f10)) — "
        "f5 used as primary gross spend anchor; f6-f10 as independent behavioral modifiers."
    ),

    "Validation Approach": (
        "1. Zero-floor file sanity check: script aborts if data_ranked.csv is the stale average-rank "
        "file (verifies mean rank of zero-value rows < 0.01 for f21, f13, f4, f15). "
        "2. Perturbation stability (30 trials each noise level): all weights perturbed ±5/10/20% randomly. "
        "Target: >=80% top-20% overlap at ±10%, >=65% at ±20%. A robust equation should not collapse "
        "under small coefficient changes. "
        "3. Ablation analysis: each feature removed individually; top-20% membership change recorded. "
        "Load-bearing (>5% change), marginal (2-5%), decorative (<2%). "
        "4. Economic gates (strict — no tolerance rigging): "
        "- 100% of top 20% have f3=0 (collections override). "
        "- Top 20% mean spend >= 1.5x bottom 80%. "
        "- Top 20% mean risk score < bottom 80%. "
        "- Top 20% mean revolve balance >= bottom 80%. "
        "- >=30% of top 20% are simultaneously high-revolve AND low-risk (ideal profitable archetype). "
        "- Benefit-inversion guard (strict): top 20% mean for f13/f14/f15/f21 each <= bottom 80% mean. "
        "All gates must pass before submitting. Tolerance bands not used — if a feature fails, "
        "the correct response is exclusion, not widening the gate."
    ),

    "Additional Notes (Optional)": (
        "Effective vs raw weight distinction: all equation coefficients are EFFECTIVE weights — "
        "the actual desired sorting power over the full 500K population. Because features with many zeros "
        "(e.g., f21 at 72% zeros) only sort the non-zero fraction of customers, the weight stated in the "
        "equation must be inflated to achieve that impact. Formula: raw = effective / coverage, where "
        "coverage = 1 - zero_fraction. Example: f21 effective -0.10, coverage 27.8%, raw = -0.360. "
        "This design makes every coefficient economically interpretable for a Round 3 presentation — "
        "the effective column is what matters, the raw column is a mechanical artifact. "
        "\n"
        "f16 excluded (not down-weighted): hard cap at $64.40 means 75th percentile = maximum value. "
        "Above the 75th percentile, every customer carries the identical cost — f16 cannot differentiate "
        "anyone in the top quartile. The highest spenders (who we want in the top 20%) also use this credit, "
        "so any weight produces an inversion in that segment. Exclusion is the correct analytical decision."
    ),
}


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — LOAD DATA
# ══════════════════════════════════════════════════════════════════════════════
section("STEP 1 — LOAD DATA")

df  = pd.read_csv(RANKED_PATH)
imp = pd.read_csv(IMPUTED_PATH)
assert (df["id"].values == imp["id"].values).all(), "ranked and imputed files not row-aligned"
print(f"  Ranked  : {len(df):,} rows")
print(f"  Imputed : {len(imp):,} rows")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — SANITY CHECK (zero-floor file)
# ══════════════════════════════════════════════════════════════════════════════
section("STEP 2 — SANITY CHECK (zero-floor file)")

stale = []
for f in ["f21", "f13", "f4", "f15"]:
    zero_rank = df.loc[imp[f] == 0, f"{f}_rank"].mean()
    ok = zero_rank < 0.01
    print(f"  {f}: zero-value rows mean rank = {zero_rank:.4f}  "
          f"{'OK' if ok else 'STALE — re-run eda_pipeline.py first'}")
    if not ok:
        stale.append(f)

if stale:
    raise ValueError(
        f"data_ranked.csv is the OLD average-rank file for {stale}. "
        "Re-run eda_pipeline.py to regenerate with zero-floor, then retry."
    )
print("  Zero-floor confirmed — effective = raw x coverage is valid.")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — BACK-SOLVE RAW WEIGHTS AND SCORE
# ══════════════════════════════════════════════════════════════════════════════
section("STEP 3 — COMPUTE SCORES")

coverage    = {f: 1.0 - (imp[f] == 0).mean() for f in EFFECTIVE_WEIGHTS}
RAW_WEIGHTS = {f: EFFECTIVE_WEIGHTS[f] / coverage[f] for f in EFFECTIVE_WEIGHTS}
WEIGHTS     = {f"{f}_rank": w for f, w in RAW_WEIGHTS.items()}

print(f"\n  {'Feat':<6} {'Coverage':>9} {'Eff W':>8} {'Raw W':>8}")
print("  " + "-" * 36)
for f in EFFECTIVE_WEIGHTS:
    print(f"  {f:<6} {coverage[f]*100:>8.1f}%  {EFFECTIVE_WEIGHTS[f]:>+8.3f}  {RAW_WEIGHTS[f]:>+8.3f}")

# weighted sum
score = sum(w * df[feat] for feat, w in WEIGHTS.items())

# f3 smooth bottom-rank override (no -999 sentinel)
ineligible  = imp["f3"].values > 0
n_inelig    = int(ineligible.sum())
eligible_min = score[~ineligible].min()
span = score.max() - score.min()
eps  = 1e-6 * (span if span > 0 else 1.0)
offset = (score[ineligible].max() - eligible_min) + eps
score = score.copy()
score.loc[ineligible] = score[ineligible] - offset

df["profit_score"] = score

print(f"\n  f3>0 customers (ineligible) : {n_inelig:,} ({n_inelig/len(df)*100:.1f}%)")
print(f"  Score range                 : [{score.min():.4f}, {score.max():.4f}]  (no -999)")
print(f"  Null scores                 : {score.isna().sum()}")
print(f"  Duplicate IDs               : {df['id'].duplicated().sum()}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — COMPLIANCE CHECKS
# ══════════════════════════════════════════════════════════════════════════════
section("STEP 4 — COMPLIANCE CHECKS")

assert len(df) == 500_000,          f"Expected 500,000 rows, got {len(df)}"
assert df["id"].nunique() == 500_000, "Duplicate IDs found"
assert not df["profit_score"].isna().any(), "Null scores present"
assert not (df["profit_score"] <= -900).any(), "-999 sentinels found"
assert "id" not in EFFECTIVE_WEIGHTS, "COMPLIANCE: identifier used in equation"

print("  [PASS]  500,000 rows")
print("  [PASS]  All IDs unique")
print("  [PASS]  No null scores")
print("  [PASS]  No -999 sentinels")
print("  [PASS]  Identifier (id) not used in equation")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — WRITE XLSX (both sheets)
# ══════════════════════════════════════════════════════════════════════════════
section("STEP 5 — WRITE FINAL SUBMISSION XLSX")

# sort by ID to align with template order
df_sorted = df[["id", "profit_score"]].sort_values("id").reset_index(drop=True)

# load the template (preserves formatting)
wb = openpyxl.load_workbook(TEMPLATE_PATH)

# ── Sheet 1: Predictions ──────────────────────────────────────────────────────
ws_pred = wb["Predictions"]
print(f"\n  Filling 'Predictions' sheet ({len(df_sorted):,} rows) ...")

for i, score_val in enumerate(df_sorted["profit_score"], start=2):  # row 1 = header
    ws_pred.cell(row=i, column=2, value=round(float(score_val), 8))

# quick spot-check
print(f"  Row 2  — ID {ws_pred.cell(2, 1).value}, score {ws_pred.cell(2, 2).value:.6f}")
print(f"  Row 501 — ID {ws_pred.cell(501, 1).value}, score {ws_pred.cell(501, 2).value:.6f}")

# ── Sheet 2: Profitability Framework ─────────────────────────────────────────
ws_fw = wb["Profitability Framework"]
print(f"\n  Filling 'Profitability Framework' sheet ...")

for row_idx in range(2, ws_fw.max_row + 1):
    section_name = ws_fw.cell(row=row_idx, column=1).value
    if section_name and section_name in FRAMEWORK:
        ws_fw.cell(row=row_idx, column=2, value=FRAMEWORK[section_name])
        print(f"    [{row_idx}] {section_name} — filled ({len(FRAMEWORK[section_name])} chars)")

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"\n  Saved --> {OUT_PATH}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 — VERIFY OUTPUT
# ══════════════════════════════════════════════════════════════════════════════
section("STEP 6 — VERIFY OUTPUT FILE")

verify = pd.read_excel(OUT_PATH, sheet_name="Predictions")
print(f"  Rows             : {len(verify):,}")
print(f"  Columns          : {list(verify.columns)}")
print(f"  Null Predictions : {verify['Prediction'].isna().sum()}")
print(f"  Unique IDs       : {verify['ID'].nunique():,}")
print(f"  Score range      : [{verify['Prediction'].min():.4f}, {verify['Prediction'].max():.4f}]")
print(f"  Top 5 rows:\n{verify.head()}")

fw_verify = pd.read_excel(OUT_PATH, sheet_name="Profitability Framework")
filled = fw_verify["Response"].notna().sum()
print(f"\n  Framework sections filled : {filled}/{len(fw_verify)}")

print(f"""
  ──────────────────────────────────────────────────
  READY TO SUBMIT: {OUT_PATH}
  Upload this file to Unstop.
  Both sheets are filled.
  ──────────────────────────────────────────────────
""")
